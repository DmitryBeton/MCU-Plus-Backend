import json
import re
from dataclasses import dataclass
from datetime import date, datetime, time

from django.core.exceptions import ValidationError
from django.db import transaction
from openpyxl import load_workbook

from .models import Institute, ScheduleEntry, StudyGroup


@dataclass
class ImportSummary:
    institutes_created: int = 0
    groups_created: int = 0
    entries_created: int = 0
    entries_updated: int = 0


def import_schedule_file(uploaded_file, default_institute_name=None):
    file_name = uploaded_file.name.lower()
    if file_name.endswith('.json'):
        return import_schedule_json(uploaded_file)
    if file_name.endswith('.xlsx'):
        return import_schedule_xlsx(uploaded_file, default_institute_name=default_institute_name)
    raise ValidationError('Поддерживаются только файлы .json и .xlsx.')


def import_schedule_json(uploaded_file):
    try:
        raw_data = uploaded_file.read().decode('utf-8-sig')
        data = json.loads(raw_data)
    except UnicodeDecodeError as error:
        raise ValidationError('Файл должен быть сохранен в UTF-8.') from error
    except json.JSONDecodeError as error:
        raise ValidationError(f'Некорректный JSON: {error.msg}.') from error

    records = list(_iter_schedule_records(data))
    if not records:
        raise ValidationError('В JSON не найдено ни одного занятия.')

    summary = ImportSummary()
    with transaction.atomic():
        for index, record in enumerate(records, start=1):
            normalized = _normalize_record(record, index)
            institute, institute_created = Institute.objects.get_or_create(
                name=normalized['institute']
            )
            group, group_created = StudyGroup.objects.get_or_create(
                institute=institute,
                course=normalized['course'],
                name=normalized['group'],
            )
            _entry, entry_created = ScheduleEntry.objects.update_or_create(
                group=group,
                date=normalized['date'],
                start_time=normalized['start_time'],
                end_time=normalized['end_time'],
                subject=normalized['subject'],
                defaults={
                    'teacher': normalized['teacher'],
                    'room': normalized['room'],
                    'note': normalized['note'],
                },
            )

            summary.institutes_created += int(institute_created)
            summary.groups_created += int(group_created)
            if entry_created:
                summary.entries_created += 1
            else:
                summary.entries_updated += 1

    return summary


def import_schedule_xlsx(uploaded_file, default_institute_name=None):
    try:
        workbook = load_workbook(uploaded_file, data_only=True)
    except Exception as error:
        raise ValidationError('Не удалось прочитать Excel-файл.') from error

    records = []
    for worksheet in workbook.worksheets:
        records.extend(_iter_excel_records(worksheet, default_institute_name))

    if not records:
        raise ValidationError('В Excel-файле не найдено ни одного занятия.')

    summary = ImportSummary()
    with transaction.atomic():
        for record in records:
            institute, institute_created = Institute.objects.get_or_create(
                name=record['institute']
            )
            group, group_created = StudyGroup.objects.get_or_create(
                institute=institute,
                course=record['course'],
                name=record['group'],
            )
            _entry, entry_created = ScheduleEntry.objects.update_or_create(
                group=group,
                date=record['date'],
                start_time=record['start_time'],
                end_time=record['end_time'],
                subject=record['subject'],
                defaults={
                    'teacher': record['teacher'],
                    'room': record['room'],
                    'note': record['note'],
                },
            )
            summary.institutes_created += int(institute_created)
            summary.groups_created += int(group_created)
            if entry_created:
                summary.entries_created += 1
            else:
                summary.entries_updated += 1

    return summary


def _iter_schedule_records(data):
    if isinstance(data, list):
        yield from data
        return

    if not isinstance(data, dict):
        raise ValidationError('Корневой элемент JSON должен быть объектом или списком.')

    if 'entries' in data:
        yield from _require_list(data['entries'], 'entries')
        return

    if 'groups' in data:
        for group in _require_list(data['groups'], 'groups'):
            for lesson in _lessons_from_group(group):
                yield {
                    **lesson,
                    'institute': _pick(group, 'institute', 'institute_name', 'институт'),
                    'course': _pick(group, 'course', 'курс'),
                    'group': _pick(group, 'group', 'group_name', 'name', 'группа'),
                }
        return

    for institute in _require_list(data.get('institutes', []), 'institutes'):
        institute_name = _pick(institute, 'name', 'institute', 'institute_name', 'институт')
        for group in _require_list(institute.get('groups', []), 'groups'):
            for lesson in _lessons_from_group(group):
                yield {
                    **lesson,
                    'institute': institute_name,
                    'course': _pick(group, 'course', 'курс'),
                    'group': _pick(group, 'name', 'group', 'group_name', 'группа'),
                }


def _lessons_from_group(group):
    lessons = (
        group.get('lessons')
        or group.get('schedule')
        or group.get('entries')
        or group.get('занятия')
        or []
    )
    yield from _require_list(lessons, 'lessons')


def _normalize_record(record, index):
    if not isinstance(record, dict):
        raise ValidationError(f'Занятие #{index}: ожидается объект.')

    normalized = {
        'institute': _required_text(record, index, 'institute', 'institute_name', 'институт'),
        'group': _required_text(record, index, 'group', 'group_name', 'группа'),
        'subject': _required_text(record, index, 'subject', 'дисциплина', 'lesson'),
        'teacher': _optional_text(record, 'teacher', 'преподаватель'),
        'room': _optional_text(record, 'room', 'auditorium', 'аудитория'),
        'note': _optional_text(record, 'note', 'comment', 'заметка'),
        'course': _parse_course(_pick(record, 'course', 'курс'), index),
        'date': _parse_date(_pick(record, 'date', 'дата'), index),
        'start_time': _parse_time(_pick(record, 'start_time', 'start', 'начало'), index, 'начало'),
        'end_time': _parse_time(_pick(record, 'end_time', 'end', 'конец'), index, 'конец'),
    }

    if normalized['end_time'] <= normalized['start_time']:
        raise ValidationError(f'Занятие #{index}: время окончания должно быть позже начала.')

    return normalized


def _pick(mapping, *keys):
    if not isinstance(mapping, dict):
        return None
    for key in keys:
        value = mapping.get(key)
        if value not in (None, ''):
            return value
    return None


def _required_text(mapping, index, *keys):
    value = _optional_text(mapping, *keys)
    if not value:
        raise ValidationError(f'Занятие #{index}: поле "{keys[0]}" обязательно.')
    return value


def _optional_text(mapping, *keys):
    value = _pick(mapping, *keys)
    if value is None:
        return ''
    return str(value).strip()


def _parse_course(value, index):
    try:
        course = int(value)
    except (TypeError, ValueError) as error:
        raise ValidationError(f'Занятие #{index}: курс должен быть числом.') from error
    if course < 1 or course > 6:
        raise ValidationError(f'Занятие #{index}: курс должен быть от 1 до 6.')
    return course


def _parse_date(value, index):
    if not value:
        raise ValidationError(f'Занятие #{index}: поле "date" обязательно.')
    text = str(value).strip()
    for parser in (date.fromisoformat, _parse_ru_date):
        try:
            return parser(text)
        except ValueError:
            continue
    raise ValidationError(f'Занятие #{index}: дата должна быть в формате YYYY-MM-DD или DD.MM.YYYY.')


def _parse_ru_date(value):
    day, month, year = value.split('.')
    return date(int(year), int(month), int(day))


def _parse_time(value, index, field_name):
    if not value:
        raise ValidationError(f'Занятие #{index}: поле "{field_name}" обязательно.')
    text = str(value).strip()
    if len(text.split(':')) == 2:
        text = f'{text}:00'
    try:
        return time.fromisoformat(text)
    except ValueError as error:
        raise ValidationError(f'Занятие #{index}: время "{field_name}" должно быть в формате HH:MM.') from error


def _require_list(value, field_name):
    if not isinstance(value, list):
        raise ValidationError(f'Поле "{field_name}" должно быть списком.')
    return value


def _iter_excel_records(worksheet, default_institute_name):
    course, group_name = _parse_sheet_group(worksheet.title)
    institute_name = _clean_text(default_institute_name) or 'Не указан'
    header_rows = [
        row
        for row in range(1, worksheet.max_row + 1)
        if 'день' in _clean_text(worksheet.cell(row=row, column=1).value).lower()
        and 'дата' in _clean_text(worksheet.cell(row=row, column=1).value).lower()
    ]

    for header_row in header_rows:
        dates_row = header_row + 1
        date_columns = _extract_date_columns(worksheet, dates_row)
        if not date_columns:
            continue

        next_header_row = min(
            [row for row in header_rows if row > header_row] or [worksheet.max_row + 1]
        )
        for row in range(dates_row + 1, next_header_row):
            time_range = _parse_excel_time_range(worksheet.cell(row=row, column=2).value)
            if not time_range:
                continue
            start_time, end_time = time_range
            for column, lesson_date in date_columns:
                lesson_text = worksheet.cell(row=row, column=column).value
                if not _clean_text(lesson_text):
                    continue
                lesson = _parse_excel_lesson_text(lesson_text)
                yield {
                    'institute': institute_name,
                    'course': course,
                    'group': group_name,
                    'date': lesson_date,
                    'start_time': start_time,
                    'end_time': end_time,
                    **lesson,
                }


def _parse_sheet_group(sheet_title):
    normalized = _clean_text(sheet_title)
    match = re.search(r'(?P<course>[1-6])\s*к\s+(?P<group>[А-ЯA-ZЁ0-9-]+)', normalized, re.IGNORECASE)
    if not match:
        raise ValidationError(
            f'Не удалось определить курс и группу из названия листа "{sheet_title}".'
        )
    return int(match.group('course')), match.group('group').upper()


def _extract_date_columns(worksheet, dates_row):
    columns = []
    for column in range(1, worksheet.max_column + 1):
        value = worksheet.cell(row=dates_row, column=column).value
        lesson_date = _excel_date(value)
        if lesson_date:
            columns.append((column, lesson_date))
    return columns


def _excel_date(value):
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date) and 2000 <= value.year <= 2100:
        return value
    return None


def _parse_excel_time_range(value):
    text = _clean_text(value)
    if not text:
        return None
    match = re.search(r'(?P<start>\d{1,2})[.:](?P<start_min>\d{2})\s*[-–]\s*(?P<end>\d{1,2})[.:](?P<end_min>\d{2})', text)
    if not match:
        return None
    start_time = time(int(match.group('start')), int(match.group('start_min')))
    end_time = time(int(match.group('end')), int(match.group('end_min')))
    if end_time <= start_time:
        return None
    return start_time, end_time


def _parse_excel_lesson_text(value):
    parts = [
        part.strip(' .')
        for part in re.split(r'\s*\n+\s*', str(value))
        if part and part.strip(' .')
    ]
    if not parts:
        return {
            'subject': 'Занятие',
            'teacher': '',
            'room': '',
            'note': '',
        }

    notes = []
    if re.fullmatch(r'\d+\s+подгруппа', parts[0], re.IGNORECASE):
        notes.append(parts.pop(0))

    subject = parts.pop(0) if parts else 'Занятие'
    room = ''
    lesson_type = ''
    teacher_parts = []

    for part in parts:
        lower = part.lower()
        if 'ауд' in lower or lower in {'дистанционно', 'онлайн'}:
            room = part
        elif part.upper() in {'ЛК', 'ПР', 'ЛБ', 'ЗАЧ', 'ЭКЗ'}:
            lesson_type = part.upper()
        else:
            teacher_parts.append(part)

    if lesson_type:
        notes.append(lesson_type)

    return {
        'subject': subject,
        'teacher': ' '.join(teacher_parts),
        'room': room,
        'note': ', '.join(notes),
    }


def _clean_text(value):
    if value is None:
        return ''
    return re.sub(r'\s+', ' ', str(value)).strip()
